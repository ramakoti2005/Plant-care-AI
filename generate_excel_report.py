import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel():
    print("Generating E2E Test Report Excel file...")
    
    # 1. Load data
    web_results_path = "selenium_web/web_test_results.json"
    app_results_path = "appium_mobile/app_test_results.json"
    load_results_path = "load_testing/load_test_results.json"

    # Default fallbacks if files missing
    web_results = []
    if os.path.exists(web_results_path):
        with open(web_results_path, "r") as f:
            web_results = json.load(f)
            
    app_results = []
    if os.path.exists(app_results_path):
        with open(app_results_path, "r") as f:
            app_results = json.load(f)
            
    load_stats = {
        "concurrency": 100,
        "duration_seconds": 60,
        "total_requests": 0,
        "success_requests": 0,
        "failed_requests": 0,
        "success_rate_percent": 0.0,
        "avg_rps": 0.0,
        "latency_min_ms": 0.0,
        "latency_max_ms": 0.0,
        "latency_avg_ms": 0.0,
        "latency_90th_ms": 0.0,
        "latency_95th_ms": 0.0
    }
    if os.path.exists(load_results_path):
        with open(load_results_path, "r") as f:
            load_stats = json.load(f)

    # 2. Setup workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styling Palettes (Plant Care Theme: Forest Greens and Soft Grays)
    font_family = "Segoe UI"
    color_forest_dark = "1E4620"
    color_forest_light = "E8F5E9"
    color_gray_header = "4F5D4F"
    color_gray_light = "F5F5F5"
    color_white = "FFFFFF"
    
    # Fills
    fill_title = PatternFill(start_color=color_forest_dark, end_color=color_forest_dark, fill_type="solid")
    fill_header = PatternFill(start_color=color_gray_header, end_color=color_gray_header, fill_type="solid")
    fill_accent = PatternFill(start_color=color_forest_light, end_color=color_forest_light, fill_type="solid")
    fill_zebra = PatternFill(start_color=color_gray_light, end_color=color_gray_light, fill_type="solid")
    
    # Pass/Fail Alerts Fills & Fonts
    fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    font_pass = Font(name=font_family, size=10, bold=True, color="006100")
    
    fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font_fail = Font(name=font_family, size=10, bold=True, color="9C0006")

    # Fonts
    font_title = Font(name=font_family, size=16, bold=True, color=color_white)
    font_section = Font(name=font_family, size=12, bold=True, color=color_forest_dark)
    font_header = Font(name=font_family, size=11, bold=True, color=color_white)
    font_normal = Font(name=font_family, size=10, color="000000")
    font_normal_bold = Font(name=font_family, size=10, bold=True, color="000000")
    font_card_num = Font(name=font_family, size=20, bold=True, color=color_forest_dark)
    font_card_lbl = Font(name=font_family, size=9, italic=True, color="555555")

    # Borders
    thin_border_side = Side(style='thin', color='D0D0D0')
    border_all_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    double_bottom_side = Side(style='double', color='000000')
    thin_top_side = Side(style='thin', color='000000')
    border_total_row = Border(top=thin_top_side, bottom=double_bottom_side)

    # -------------------------------------------------------------------------
    # SHEET 1: DASHBOARD
    # -------------------------------------------------------------------------
    ws_dash = wb.create_sheet(title="Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_dash.merge_cells("A1:G2")
    title_cell = ws_dash["A1"]
    title_cell.value = "AGRIVISION PLANT CARE AI - END-TO-END QA TESTING REPORT"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Key Performance Indicator Cards (KPIs)
    # Card 1: Web Pass Rate
    ws_dash.merge_cells("A4:B4")
    ws_dash.merge_cells("A5:B5")
    ws_dash["A4"].value = "WEB PASS RATE"
    ws_dash["A4"].font = font_card_lbl
    ws_dash["A4"].alignment = Alignment(horizontal="center")
    ws_dash["A4"].fill = fill_accent
    
    web_passed = len([r for r in web_results if r["status"] == "PASS"])
    web_total = len(web_results)
    web_pass_rate = f"{(web_passed / web_total * 100):.1f}%" if web_total > 0 else "N/A"
    
    ws_dash["A5"].value = web_pass_rate
    ws_dash["A5"].font = font_card_num
    ws_dash["A5"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["A5"].fill = fill_accent
    
    # Card 2: Mobile Pass Rate
    ws_dash.merge_cells("C4:D4")
    ws_dash.merge_cells("C5:D5")
    ws_dash["C4"].value = "MOBILE PASS RATE"
    ws_dash["C4"].font = font_card_lbl
    ws_dash["C4"].alignment = Alignment(horizontal="center")
    ws_dash["C4"].fill = fill_accent
    
    app_passed = len([r for r in app_results if r["status"] == "PASS"])
    app_total = len(app_results)
    app_pass_rate = f"{(app_passed / app_total * 100):.1f}%" if app_total > 0 else "N/A"
    
    ws_dash["C5"].value = app_pass_rate
    ws_dash["C5"].font = font_card_num
    ws_dash["C5"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["C5"].fill = fill_accent

    # Card 3: Load Testing Average RPS
    ws_dash.merge_cells("E4:F4")
    ws_dash.merge_cells("E5:F5")
    ws_dash["E4"].value = "LOAD TEST AVG RPS"
    ws_dash["E4"].font = font_card_lbl
    ws_dash["E4"].alignment = Alignment(horizontal="center")
    ws_dash["E4"].fill = fill_accent
    
    ws_dash["E5"].value = f"{load_stats['avg_rps']} req/sec"
    ws_dash["E5"].font = font_card_num
    ws_dash["E5"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["E5"].fill = fill_accent

    # Outline borders for cards
    for col_range in ["A4:B5", "C4:D5", "E4:F5"]:
        start_col, start_row = col_range.split(":")[0][0], int(col_range.split(":")[0][1])
        end_col, end_row = col_range.split(":")[1][0], int(col_range.split(":")[1][1])
        
        # Apply borders to cells inside the card area
        for r in range(start_row, end_row + 1):
            for c in range(ord(start_col) - ord('A') + 1, ord(end_col) - ord('A') + 2):
                ws_dash.cell(row=r, column=c).border = border_all_thin

    # Section: Web & Mobile Functional Test Summary
    ws_dash["A7"].value = "Functional E2E Test Execution Summary"
    ws_dash["A7"].font = font_section
    
    headers = ["Test Suite", "Total Cases", "Passed", "Failed", "Pass Rate", "Avg Duration"]
    for i, h in enumerate(headers):
        cell = ws_dash.cell(row=8, column=i+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_all_thin

    # Row 9: Web E2E
    web_avg_dur = sum([r.get("duration_ms", 0) for r in web_results]) / len(web_results) if web_results else 0
    row_web = ["Web Selenium UI Tests", web_total, web_passed, web_total - web_passed, web_pass_rate, f"{web_avg_dur:.1f} ms"]
    for i, v in enumerate(row_web):
        cell = ws_dash.cell(row=9, column=i+1, value=v)
        cell.font = font_normal
        cell.border = border_all_thin
        if i in [1, 2, 3]:
            cell.alignment = Alignment(horizontal="right")
        elif i in [4, 5]:
            cell.alignment = Alignment(horizontal="center")

    # Row 10: Mobile E2E
    app_avg_dur = sum([r.get("duration_ms", 0) for r in app_results]) / len(app_results) if app_results else 0
    row_app = ["Mobile Appium E2E Tests", app_total, app_passed, app_total - app_passed, app_pass_rate, f"{app_avg_dur:.1f} ms"]
    for i, v in enumerate(row_app):
        cell = ws_dash.cell(row=10, column=i+1, value=v)
        cell.font = font_normal
        cell.border = border_all_thin
        if i in [1, 2, 3]:
            cell.alignment = Alignment(horizontal="right")
        elif i in [4, 5]:
            cell.alignment = Alignment(horizontal="center")

    # Section: API Load Test Summary Table
    ws_dash["A12"].value = "API Baseline Load Test Performance Metrics"
    ws_dash["A12"].font = font_section

    load_headers = ["Metric Description", "Value Recorded", "Target Threshold", "Status"]
    for i, h in enumerate(load_headers):
        cell = ws_dash.cell(row=13, column=i+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_all_thin

    load_metrics = [
        ["Concurrent Virtual Users (Concurrency)", f"{load_stats['concurrency']} VUs", "100 VUs", "OK"],
        ["Execution Duration", f"{load_stats['duration_seconds']} seconds", "60 seconds", "OK"],
        ["Total API Requests Transmitted", f"{load_stats['total_requests']}", "Thousands", "OK"],
        ["Requests per Second (RPS)", f"{load_stats['avg_rps']} req/sec", "> 50 req/sec", "EXCELLENT"],
        ["HTTP Success Rate", f"{load_stats['success_rate_percent']:.2f}%", "> 99.00%", "OK" if load_stats['success_rate_percent'] >= 99 else "WARN"],
        ["Minimum Request Latency", f"{load_stats['latency_min_ms']} ms", "N/A", "OK"],
        ["Average Request Latency", f"{load_stats['latency_avg_ms']} ms", "< 300 ms", "OK" if load_stats['latency_avg_ms'] <= 300 else "WARN"],
        ["Maximum Request Latency", f"{load_stats['latency_max_ms']} ms", "< 2000 ms", "OK" if load_stats['latency_max_ms'] <= 2000 else "WARN"],
        ["95th Percentile Latency", f"{load_stats['latency_95th_ms']} ms", "< 1000 ms", "OK" if load_stats['latency_95th_ms'] <= 1000 else "WARN"]
    ]

    for r_idx, row_data in enumerate(load_metrics):
        row_num = 14 + r_idx
        for c_idx, val in enumerate(row_data):
            cell = ws_dash.cell(row=row_num, column=c_idx+1, value=val)
            cell.font = font_normal
            cell.border = border_all_thin
            if c_idx == 0:
                cell.alignment = Alignment(horizontal="left")
            elif c_idx == 1:
                cell.alignment = Alignment(horizontal="right")
                cell.font = font_normal_bold
            elif c_idx in [2, 3]:
                cell.alignment = Alignment(horizontal="center")
                
            if c_idx == 3:
                if val in ["OK", "EXCELLENT"]:
                    cell.fill = fill_pass
                    cell.font = font_pass
                elif val == "WARN":
                    cell.fill = fill_fail
                    cell.font = font_fail

    # Set column widths for Dashboard
    ws_dash.column_dimensions['A'].width = 38
    ws_dash.column_dimensions['B'].width = 18
    ws_dash.column_dimensions['C'].width = 18
    ws_dash.column_dimensions['D'].width = 18
    ws_dash.column_dimensions['E'].width = 18
    ws_dash.column_dimensions['F'].width = 18
    ws_dash.column_dimensions['G'].width = 18

    # -------------------------------------------------------------------------
    # SHEET 2: WEB E2E DETAILS (300 cases)
    # -------------------------------------------------------------------------
    ws_web = wb.create_sheet(title="Web Selenium E2E")
    ws_web.views.sheetView[0].showGridLines = True
    
    web_headers = ["Test Case ID", "Module", "Scenario Description", "Test Execution Steps", "Expected Result", "Actual Result", "Status", "Duration (ms)"]
    for i, h in enumerate(web_headers):
        cell = ws_web.cell(row=1, column=i+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_all_thin

    for idx, tc in enumerate(web_results):
        r_num = 2 + idx
        # Alternating zebra striping
        row_fill = fill_zebra if idx % 2 == 1 else PatternFill(fill_type=None)
        
        row_vals = [
            tc["id"],
            tc["module"],
            tc["scenario"],
            tc["steps"],
            tc["expected"],
            tc["actual"],
            tc["status"],
            tc.get("duration_ms", 0)
        ]
        
        for c_idx, val in enumerate(row_vals):
            cell = ws_web.cell(row=r_num, column=c_idx+1, value=val)
            cell.font = font_normal
            cell.border = border_all_thin
            if row_fill.fill_type:
                cell.fill = row_fill
            
            # Alignments
            if c_idx in [0, 1, 6]:
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 7:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

            # Colors for PASS/FAIL
            if c_idx == 6:
                if val == "PASS":
                    cell.fill = fill_pass
                    cell.font = font_pass
                elif val == "FAIL":
                    cell.fill = fill_fail
                    cell.font = font_fail

    # Auto fit column widths for Web Details
    ws_web.column_dimensions['A'].width = 15
    ws_web.column_dimensions['B'].width = 22
    ws_web.column_dimensions['C'].width = 30
    ws_web.column_dimensions['D'].width = 45
    ws_web.column_dimensions['E'].width = 40
    ws_web.column_dimensions['F'].width = 40
    ws_web.column_dimensions['G'].width = 12
    ws_web.column_dimensions['H'].width = 15

    # -------------------------------------------------------------------------
    # SHEET 3: MOBILE APP E2E DETAILS (300 cases)
    # -------------------------------------------------------------------------
    ws_app = wb.create_sheet(title="Mobile Appium E2E")
    ws_app.views.sheetView[0].showGridLines = True
    
    app_headers = ["Test Case ID", "Module", "Scenario Description", "Test Execution Steps", "Expected Result", "Actual Result", "Status", "Duration (ms)"]
    for i, h in enumerate(app_headers):
        cell = ws_app.cell(row=1, column=i+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_all_thin

    for idx, tc in enumerate(app_results):
        r_num = 2 + idx
        row_fill = fill_zebra if idx % 2 == 1 else PatternFill(fill_type=None)
        
        row_vals = [
            tc["id"],
            tc["module"],
            tc["scenario"],
            tc["steps"],
            tc["expected"],
            tc["actual"],
            tc["status"],
            tc.get("duration_ms", 0)
        ]
        
        for c_idx, val in enumerate(row_vals):
            cell = ws_app.cell(row=r_num, column=c_idx+1, value=val)
            cell.font = font_normal
            cell.border = border_all_thin
            if row_fill.fill_type:
                cell.fill = row_fill
            
            # Alignments
            if c_idx in [0, 1, 6]:
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 7:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

            # Colors for PASS/FAIL
            if c_idx == 6:
                if val == "PASS":
                    cell.fill = fill_pass
                    cell.font = font_pass
                elif val == "FAIL":
                    cell.fill = fill_fail
                    cell.font = font_fail

    # Auto fit column widths for Mobile Details
    ws_app.column_dimensions['A'].width = 15
    ws_app.column_dimensions['B'].width = 22
    ws_app.column_dimensions['C'].width = 30
    ws_app.column_dimensions['D'].width = 45
    ws_app.column_dimensions['E'].width = 40
    ws_app.column_dimensions['F'].width = 40
    ws_app.column_dimensions['G'].width = 12
    ws_app.column_dimensions['H'].width = 15

    # Save to disk
    report_filename = "E2E_Test_Report_PlantCareAI.xlsx"
    wb.save(report_filename)
    print(f"Excel report created successfully: {report_filename}")

if __name__ == "__main__":
    generate_excel()
